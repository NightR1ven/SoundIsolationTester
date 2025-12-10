# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import time
import sys
import os
import json
import webbrowser
from datetime import datetime, timedelta
import csv
import subprocess
import random
import math

# Функция для динамического импорта модулей
def import_audio_core():
    """Динамический импорт AudioCore"""
    try:
        from audio_core import AudioCore
        return AudioCore
    except ImportError as e:
        print(f"⚠️ Ошибка импорта AudioCore: {e}")
        # Создаем заглушку
        class AudioCoreStub:
            def __init__(self):
                self.is_recording = False
                print("⚠️ Используется заглушка AudioCore")
            
            def get_audio_devices(self):
                return []
            
            def start_recording(self, *args, **kwargs):
                print("⚠️ Заглушка: запись невозможна")
                return False
            
            def stop_recording(self):
                return {}
            
            def get_recording_stats(self):
                return {}
            
            def get_audio_levels(self):
                return {'outside': 0.0, 'inside': 0.0}
            
            def cleanup(self):
                pass
        
        return AudioCoreStub

def import_ai_analyzer():
    """Динамический импорт анализатора"""
    try:
        from ai_analyzer import EnhancedSoundIsolationAnalyzer
        return EnhancedSoundIsolationAnalyzer
    except ImportError as e:
        print(f"⚠️ Ошибка импорта анализатора: {e}")
        # Создаем заглушку
        class AnalyzerStub:
            def __init__(self):
                self.initialized = False
                
            def analyze_with_audio_analysis(self, *args, **kwargs):
                return {'results': {'overall_assessment': {'verdict': 'УСТАНОВИТЕ ЗАВИСИМОСТИ'}}}
            
            def set_recognition_engine(self, engine_name):
                return False
        
        return AnalyzerStub

def import_speech_recognizer():
    """Динамический импорт распознавателя речи"""
    try:
        # Проверяем, есть ли файл
        if not os.path.exists("speech_recognizer.py"):
            print("⚠️ Файл speech_recognizer.py не найден")
            raise ImportError("Файл не найден")
        
        # Добавляем текущую директорию в путь
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        
        from speech_recognizer import MultiEngineSpeechRecognizer, RecognitionEngine
        return MultiEngineSpeechRecognizer, RecognitionEngine, True
    except ImportError as e:
        print(f"⚠️ Ошибка импорта распознавателя: {e}")
        # Создаем заглушку
        from enum import Enum
        
        class RecognitionEngineStub(Enum):
            WHISPER_TINY = "whisper-tiny"
            WHISPER_BASE = "whisper-base"
            WHISPER_SMALL = "whisper-small"
            WHISPER_MEDIUM = "whisper-medium"
            VOSK_SMALL_RU = "vosk-small-ru"
            VOSK_LARGE_RU = "vosk-large-ru"
        
        class RecognizerStub:
            def __init__(self, models_dir="models"):
                self.supported_engines = [
                    RecognitionEngineStub.WHISPER_TINY,
                    RecognitionEngineStub.WHISPER_SMALL,
                    RecognitionEngineStub.VOSK_SMALL_RU,
                ]
                self.current_engine = None
            
            def set_engine(self, engine):
                print(f"⚠️ Заглушка: установка движка {engine}")
                self.current_engine = engine
                return False
            
            def transcribe(self, *args, **kwargs):
                print("⚠️ Заглушка: распознавание недоступно")
                return None
            
            def analyze_pair(self, *args, **kwargs):
                return {
                    'outside': {'text': 'Распознавание недоступно. Установите модели.'}, 
                    'inside': {'text': 'Распознавание недоступно. Установите модели.'},
                    'comparison': {'wer': 1.0},
                    'engine': 'stub'
                }
            
            def calculate_wer(self, *args, **kwargs):
                return 1.0
        
        return RecognizerStub, RecognitionEngineStub, False

# Импортируем модули
AudioCore = import_audio_core()
EnhancedSoundIsolationAnalyzer = import_ai_analyzer()
MultiEngineSpeechRecognizer, RecognitionEngine, SPEECH_RECOGNITION_AVAILABLE = import_speech_recognizer()

# Пытаемся импортировать polars, если нет - используем альтернативы
try:
    import polars as pl
    POLARS_AVAILABLE = True
    print("✅ Polars загружен")
except ImportError:
    POLARS_AVAILABLE = False
    print("⚠️ Polars не установлен, используем CSV")

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass

sys.path.append(os.path.dirname(__file__))

class RecordingIndicator(tk.Canvas):
    """Анимированный индикатор записи с барами"""
    
    def __init__(self, parent, width=500, height=120, label="", **kwargs):
        super().__init__(parent, width=width, height=height, **kwargs)
        self.width = width
        self.height = height
        self.label = label
        self.level = 0.0
        self.is_active = False
        self.bars = []
        self.animation_id = None
        
        # Темный фон
        self.create_rectangle(0, 0, width, height, fill="#1a1a2e", outline="")
        
        # Заголовок
        self.title = self.create_text(
            width//2, 20, 
            text=label, 
            font=('Arial', 12, 'bold'),
            fill="white"
        )
        
        # Область для баров
        self.bar_area = self.create_rectangle(
            20, 40, width-20, height-15,
            fill="#16213e", outline="#0f3460", width=2
        )
        
        # Создаем бары (вертикальные полоски)
        bar_width = 10
        bar_spacing = 3
        num_bars = (width - 40) // (bar_width + bar_spacing)
        start_x = 25
        
        for i in range(num_bars):
            x1 = start_x + i * (bar_width + bar_spacing)
            x2 = x1 + bar_width
            bar = self.create_rectangle(
                x1, height-20, x2, height-20,  # Начинаем с минимальной высоты
                fill="#00b894", outline="#00b894"
            )
            self.bars.append(bar)
        
        # Индикатор записи (красный кружок)
        self.record_indicator = self.create_oval(
            width-35, 15, width-20, 30,
            fill="#e74c3c", outline=""
        )
        
        # Текст REC
        self.record_text = self.create_text(
            width-27, 22,
            text="●",
            font=('Arial', 10, 'bold'),
            fill="white"
        )
        
        # Текущий уровень
        self.level_text = self.create_text(
            width//2, height-5,
            text="Уровень: 0%",
            font=('Arial', 9),
            fill="#95a5a6"
        )
    
    def set_active(self, active):
        """Активировать/деактивировать индикатор"""
        self.is_active = active
        if active:
            self.itemconfig(self.record_indicator, fill="#e74c3c")
            self.itemconfig(self.record_text, text="●")
            self._start_animation()
        else:
            self.itemconfig(self.record_indicator, fill="#7f8c8d")
            self.itemconfig(self.record_text, text="○")
            self._stop_animation()
    
    def update_level(self, level):
        """Обновить уровень звука (0.0 - 1.0)"""
        self.level = max(0.0, min(1.0, level))
        
        # Обновляем текст уровня
        self.itemconfig(self.level_text, text=f"Уровень: {int(self.level*100)}%")
    
    def _start_animation(self):
        """Запустить анимацию баров"""
        if self.animation_id:
            self.after_cancel(self.animation_id)
        
        if self.is_active:
            self._animate_bars()
    
    def _stop_animation(self):
        """Остановить анимацию"""
        if self.animation_id:
            self.after_cancel(self.animation_id)
            self.animation_id = None
        
        # Сбрасываем все бары
        for bar in self.bars:
            self.coords(bar, self.coords(bar)[0], self.height-20, 
                       self.coords(bar)[2], self.height-20)
            self.itemconfig(bar, fill="#00b894")
    
    def _animate_bars(self):
        """Анимировать бары в зависимости от уровня звука"""
        if not self.is_active:
            return
        
        num_bars = len(self.bars)
        active_bars = int(self.level * num_bars)
        
        # Максимальная высота бара
        max_bar_height = self.height - 60
        
        for i, bar in enumerate(self.bars):
            # Текущие координаты бара
            x1, y1, x2, y2 = self.coords(bar)
            
            if i < active_bars:
                # Этот бар должен быть активным
                target_height = max_bar_height * (i / num_bars) + random.uniform(10, 30)
                target_top = self.height - 20 - target_height
                
                # Добавляем немного случайности для естественного вида
                target_top += random.uniform(-5, 5)
                target_top = max(self.height - 20 - max_bar_height, min(self.height - 25, target_top))
                
                # Плавная анимация к целевой позиции
                current_top = y1
                if abs(current_top - target_top) > 2:
                    new_top = current_top + (target_top - current_top) * 0.3
                else:
                    new_top = target_top
                
                # Обновляем координаты
                self.coords(bar, x1, new_top, x2, self.height-20)
                
                # Цвет в зависимости от высоты
                bar_height = (self.height - 20 - new_top) / max_bar_height
                if bar_height < 0.3:
                    color = "#00b894"  # Зеленый
                elif bar_height < 0.7:
                    color = "#fdcb6e"  # Желтый
                else:
                    color = "#e17055"  # Оранжевый/Красный
                
                self.itemconfig(bar, fill=color, outline=color)
            else:
                # Неактивный бар - опускаем вниз
                current_top = y1
                target_top = self.height - 20
                if current_top < target_top - 1:
                    new_top = current_top + (target_top - current_top) * 0.5
                    self.coords(bar, x1, new_top, x2, self.height-20)
                else:
                    self.coords(bar, x1, self.height-20, x2, self.height-20)
                    self.itemconfig(bar, fill="#00b894", outline="#00b894")
        
        # Продолжаем анимацию
        self.animation_id = self.after(50, self._animate_bars)
    
    def reset(self):
        """Сбросить индикатор"""
        self.level = 0.0
        self.set_active(False)
        self.itemconfig(self.level_text, text="Уровень: 0%")

class AdvancedSoundTester:
    def __init__(self, root):
        self.root = root
        self.root.title("Sound Isolation Tester v3.13 - Дипломная работа")
        self.root.geometry("1200x900")
        
        self.center_window()
        
        # Инициализация
        try:
            # Создаем папки
            self._create_directories()
            
            # Создаем экземпляры классов
            self.audio_core = AudioCore()
            self.analyzer = EnhancedSoundIsolationAnalyzer()
            
            # Инициализация распознавателя
            self.recognizer = None
            if SPEECH_RECOGNITION_AVAILABLE:
                try:
                    self.recognizer = MultiEngineSpeechRecognizer(models_dir="models")
                    self.current_engine = None
                    print("✅ Распознаватель речи инициализирован")
                except Exception as e:
                    print(f"⚠️ Ошибка инициализации распознавателя: {e}")
                    self.recognizer = None
            else:
                print("⚠️ Распознавание речи отключено")
            
            self.recordings_folder = "recordings"
            
            # Создаем папку для записей если не существует
            if not os.path.exists(self.recordings_folder):
                os.makedirs(self.recordings_folder)
            
            self.setup_styles()
            self.setup_ui()
            self.refresh_devices()
            self.refresh_recordings_list()
            
            # Загружаем последнюю конфигурацию
            self.load_config()
            
            # Флаг для мониторинга
            self.monitoring_active = False
            
            # Флаг истечения времени записи
            self.recording_timer_active = False
            
            print("✅ Приложение успешно инициализировано")
            
        except Exception as e:
            print(f"❌ Критическая ошибка инициализации: {e}")
            import traceback
            traceback.print_exc()
            
            error_msg = f"Ошибка инициализации:\n\n{str(e)}\n\n"
            error_msg += "Проверьте:\n"
            error_msg += "1. Все файлы находятся в одной папке\n"
            error_msg += "2. Установлены все зависимости (pip install -r requirements.txt)\n"
            error_msg += "3. Для Windows: установлен Microsoft Visual C++ Redistributable"
            
            messagebox.showerror("Ошибка инициализации", error_msg)
    
    def _create_directories(self):
        """Создание необходимых папок"""
        folders = ["models", "models/whisper", "models/vosk", "recordings", "experiments"]
        for folder in folders:
            os.makedirs(folder, exist_ok=True)
    
    def center_window(self):
        """Центрирование окна"""
        self.root.update_idletasks()
        width = 1200
        height = 900
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_styles(self):
        """Настройка стилей"""
        style = ttk.Style()
        style.configure("Red.TButton", foreground="red", font=('Arial', 10, 'bold'))
        style.configure("Green.TButton", foreground="green", font=('Arial', 10, 'bold'))
        style.configure("Title.TLabel", font=('Arial', 14, 'bold'))
    
    def setup_ui(self):
        """Настройка интерфейса"""
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Заголовок
        title = ttk.Label(main_frame, 
            text="🧪 ТЕСТЕР ЗВУКОИЗОЛЯЦИИ - Дипломная работа",
            font=('Arial', 14, 'bold'))
        title.pack(pady=10)
        
        # Вкладки
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Вкладка 1: Запись
        record_frame = ttk.Frame(notebook, padding="10")
        notebook.add(record_frame, text="🎙️ ЗАПИСЬ")
        self.setup_record_tab(record_frame)
        
        # Вкладка 2: Анализ
        analysis_frame = ttk.Frame(notebook, padding="10")
        notebook.add(analysis_frame, text="📊 АНАЛИЗ")
        self.setup_analysis_tab(analysis_frame)
        
        # Вкладка 3: Настройки движков
        engine_frame = ttk.Frame(notebook, padding="10")
        notebook.add(engine_frame, text="⚙️ ДВИЖКИ")
        self.setup_engine_tab(engine_frame)
        
        # Вкладка 4: Экспорт
        export_frame = ttk.Frame(notebook, padding="10")
        notebook.add(export_frame, text="📁 ЭКСПОРТ")
        self.setup_export_tab(export_frame)
        
        # Статус
        self.status_var = tk.StringVar(value="✅ Готов к работе")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, 
                              relief=tk.SUNKEN, padding=5)
        status_bar.pack(fill=tk.X, pady=5)
    
    def setup_record_tab(self, parent):
        """Вкладка записи"""
        # Блок 1: Устройства
        device_frame = ttk.LabelFrame(parent, text="Аудиоустройства", padding="10")
        device_frame.pack(fill=tk.X, pady=10)
        
        # Внешний микрофон
        ttk.Label(device_frame, text="Снаружи:", font=('Arial', 10)).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.outside_combo = ttk.Combobox(device_frame, width=60, state="readonly")
        self.outside_combo.grid(row=0, column=1, padx=10, pady=5)
        
        # Внутренний микрофон
        ttk.Label(device_frame, text="Внутри:", font=('Arial', 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.inside_combo = ttk.Combobox(device_frame, width=60, state="readonly")
        self.inside_combo.grid(row=1, column=1, padx=10, pady=5)
        
        # Кнопки управления устройствами
        btn_frame = ttk.Frame(device_frame)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(btn_frame, text="🔄 Обновить", command=self.refresh_devices).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🎧 Тест", command=self.test_devices).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📊 Сводка", command=self.show_device_summary).pack(side=tk.LEFT, padx=5)
        
        device_frame.columnconfigure(1, weight=1)
        
        # Блок 2: Индикаторы записи (появляются при записи)
        self.indicator_frame = ttk.LabelFrame(parent, text="Индикаторы записи", padding="10")
        self.indicator_frame.pack(fill=tk.X, pady=10)
        
        # Скрываем индикаторы по умолчанию
        self.indicator_frame.pack_forget()
        
        # Контейнер для индикаторов
        indicator_container = ttk.Frame(self.indicator_frame)
        indicator_container.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Индикатор внешнего микрофона
        self.outside_indicator = RecordingIndicator(
            indicator_container, 
            width=550, 
            height=130, 
            label="🎤 ВНЕШНИЙ МИКРОФОН (СНАРУЖИ)"
        )
        self.outside_indicator.pack(side=tk.LEFT, padx=10, fill=tk.BOTH, expand=True)
        
        # Индикатор внутреннего микрофона
        self.inside_indicator = RecordingIndicator(
            indicator_container, 
            width=550, 
            height=130, 
            label="🎤 ВНУТРЕННИЙ МИКРОФОН (ВНУТРИ)"
        )
        self.inside_indicator.pack(side=tk.LEFT, padx=10, fill=tk.BOTH, expand=True)
        
        # Блок 3: Параметры
        params_frame = ttk.LabelFrame(parent, text="Параметры теста", padding="10")
        params_frame.pack(fill=tk.X, pady=10)
        
        # Название теста
        ttk.Label(params_frame, text="Название:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.test_name_var = tk.StringVar(value=f"test_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        ttk.Entry(params_frame, textvariable=self.test_name_var, width=40).grid(row=0, column=1, padx=10, pady=5)
        
        # Длительность
        ttk.Label(params_frame, text="Длительность (сек):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.duration_var = tk.StringVar(value="15")
        ttk.Spinbox(params_frame, from_=5, to=300, textvariable=self.duration_var, width=15).grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        
        # Опции
        self.enable_analysis_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(params_frame, text="Автоматический анализ", variable=self.enable_analysis_var).grid(row=2, column=0, columnspan=2, pady=5, sticky=tk.W)
        
        # Блок 4: Управление
        control_frame = ttk.LabelFrame(parent, text="Управление записью", padding="10")
        control_frame.pack(fill=tk.X, pady=10)
        
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(expand=True)
        
        self.record_btn = ttk.Button(btn_frame, text="🔴 НАЧАТЬ ЗАПИСЬ", 
                                    command=self.start_recording,
                                    style="Red.TButton",
                                    width=20)
        self.record_btn.pack(side=tk.LEFT, padx=10, pady=10)
        
        self.stop_btn = ttk.Button(btn_frame, text="⏹️ ОСТАНОВИТЬ", 
                                  command=self.stop_recording,
                                  state=tk.DISABLED,
                                  width=20)
        self.stop_btn.pack(side=tk.LEFT, padx=10, pady=10)
        
        # Индикаторы состояния
        indicator_frame = ttk.Frame(parent)
        indicator_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(indicator_frame, text="Статус:").pack(side=tk.LEFT)
        self.record_status = ttk.Label(indicator_frame, text="Ожидание", foreground="blue")
        self.record_status.pack(side=tk.LEFT, padx=10)
        
        # Таймер
        self.timer_label = ttk.Label(indicator_frame, text="00:00 / 00:00", font=('Arial', 12, 'bold'))
        self.timer_label.pack(side=tk.RIGHT)
    
    def refresh_devices(self):
        """Обновить список аудиоустройств"""
        try:
            self.status_var.set("🔄 Поиск аудиоустройств...")
            self.root.update()
            
            devices = self.audio_core.get_audio_devices()
            
            # Очищаем комбобоксы
            self.outside_combo.set('')
            self.inside_combo.set('')
            
            # Заполняем устройствами
            device_names = [f"{i}: {d['name']} ({d['channels']} каналов)" for i, d in enumerate(devices)]
            
            self.outside_combo['values'] = device_names
            self.inside_combo['values'] = device_names
            
            if device_names:
                if len(device_names) >= 1:
                    self.outside_combo.current(0)
                if len(device_names) >= 2:
                    self.inside_combo.current(1)
                
                self.status_var.set(f"✅ Найдено устройств: {len(devices)}")
                return True
            else:
                self.status_var.set("⚠️ Устройства не найдены")
                return False
                
        except Exception as e:
            error_msg = f"❌ Ошибка обновления устройств: {e}"
            self.status_var.set(error_msg)
            messagebox.showerror("Ошибка", error_msg)
            return False
    
    def test_devices(self):
        """Тест выбранных устройств"""
        try:
            outside_idx = self.outside_combo.current()
            inside_idx = self.inside_combo.current()
            
            if outside_idx < 0 or inside_idx < 0:
                messagebox.showwarning("Предупреждение", "Выберите оба устройства")
                return
            
            # Показываем индикаторы для теста
            self.show_indicators()
            self.outside_indicator.set_active(True)
            self.inside_indicator.set_active(True)
            
            # Тестовая анимация
            self._start_test_animation()
            
            # Создаем тестовый поток для проверки устройств
            test_thread = threading.Thread(target=self._perform_device_test, 
                                          args=(outside_idx, inside_idx))
            test_thread.daemon = True
            test_thread.start()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка тестирования: {e}")
            self.hide_indicators()
    
    def _perform_device_test(self, outside_idx, inside_idx):
        """Выполнить тест устройств"""
        try:
            self.status_var.set("🔊 Тестирование устройств...")
            
            # Простая проверка - запись на 2 секунды
            success = self.audio_core.start_recording(
                outside_idx, inside_idx, duration=2, 
                test_name=f"device_test_{datetime.now().strftime('%H%M%S')}"
            )
            
            if success:
                time.sleep(2.5)  # Ждем завершения
                stats = self.audio_core.get_recording_stats()
                
                # Выключаем индикаторы
                self.root.after(0, self.hide_indicators)
                self.root.after(0, lambda: self.outside_indicator.set_active(False))
                self.root.after(0, lambda: self.inside_indicator.set_active(False))
                
                self.status_var.set(f"✅ Тест завершен. Записано: {stats.get('duration', 0):.1f} сек")
                messagebox.showinfo("Тест устройств", 
                                  f"Тест завершен успешно!\n"
                                  f"Снаружи: {stats.get('outside_samples', 0)} сэмплов\n"
                                  f"Внутри: {stats.get('inside_samples', 0)} сэмплов")
            else:
                self.root.after(0, self.hide_indicators)
                self.status_var.set("❌ Ошибка тестирования")
                messagebox.showerror("Ошибка", "Не удалось начать запись")
                
        except Exception as e:
            self.root.after(0, self.hide_indicators)
            self.status_var.set("❌ Ошибка тестирования")
            messagebox.showerror("Ошибка", f"Ошибка теста: {e}")
    
    def _start_test_animation(self):
        """Запустить тестовую анимацию индикаторов"""
        test_start = time.time()
        
        def animate():
            elapsed = time.time() - test_start
            if elapsed < 3:  # 3 секунды анимации
                # Синусоидальная анимация
                level = (math.sin(elapsed * 5) + 1) / 2
                self.outside_indicator.update_level(level * 0.8)
                self.inside_indicator.update_level(level * 0.6)
                self.root.after(50, animate)
            else:
                # Сбрасываем анимацию
                self.outside_indicator.update_level(0)
                self.inside_indicator.update_level(0)
        
        animate()
    
    def show_indicators(self):
        """Показать индикаторы записи"""
        self.indicator_frame.pack(fill=tk.X, pady=10)
        self.root.update()
    
    def hide_indicators(self):
        """Скрыть индикаторы записи"""
        self.indicator_frame.pack_forget()
        self.root.update()
    
    def start_recording(self):
        """Начать запись"""
        try:
            outside_idx = self.outside_combo.current()
            inside_idx = self.inside_combo.current()
            
            if outside_idx < 0 or inside_idx < 0:
                messagebox.showwarning("Предупреждение", "Выберите оба устройства")
                return
            
            test_name = self.test_name_var.get()
            duration = int(self.duration_var.get())
            
            # Обновляем интерфейс
            self.record_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
            self.record_status.config(text="🔴 ИДЕТ ЗАПИСЬ", foreground="red")
            self.status_var.set("🎙️ Запись начата...")
            
            # Показываем и активируем индикаторы
            self.show_indicators()
            self.outside_indicator.set_active(True)
            self.inside_indicator.set_active(True)
            
            # Запускаем мониторинг уровней
            self.monitoring_active = True
            self._start_level_monitoring()
            
            # Запускаем запись в отдельном потоке
            self.recording_thread = threading.Thread(
                target=self._perform_recording,
                args=(outside_idx, inside_idx, duration, test_name)
            )
            self.recording_thread.daemon = True
            self.recording_thread.start()
            
            # Запускаем таймер с указанием общей длительности
            self.start_time = time.time()
            self.recording_duration = duration  # Сохраняем длительность
            self._update_timer()
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка начала записи: {e}")
            self._stop_recording_ui()
    
    def _start_level_monitoring(self):
        """Запустить мониторинг уровней звука"""
        if self.monitoring_active:
            try:
                # Получаем реальные уровни звука
                levels = self.audio_core.get_audio_levels()
                
                # Обновляем индикаторы
                self.outside_indicator.update_level(levels['outside'])
                self.inside_indicator.update_level(levels['inside'])
                
                # Добавляем немного случайности для естественности
                if random.random() < 0.3:  # 30% chance
                    outside_noise = levels['outside'] + random.uniform(-0.05, 0.1)
                    inside_noise = levels['inside'] + random.uniform(-0.03, 0.07)
                    self.outside_indicator.update_level(max(0, min(1, outside_noise)))
                    self.inside_indicator.update_level(max(0, min(1, inside_noise)))
                
            except Exception as e:
                # Если ошибка, используем демо-анимацию
                demo_level = (math.sin(time.time() * 3) + 1) / 2
                self.outside_indicator.update_level(demo_level * 0.8)
                self.inside_indicator.update_level(demo_level * 0.5)
            
            # Продолжаем мониторинг
            self.root.after(100, self._start_level_monitoring)
    
    def _perform_recording(self, outside_idx, inside_idx, duration, test_name):
        """Выполнить запись"""
        try:
            # Сохраняем длительность для проверки
            self.recording_duration = duration
            
            success = self.audio_core.start_recording(
                outside_idx, inside_idx, duration, test_name
            )
            
            if not success:
                self.root.after(0, lambda: messagebox.showerror("Ошибка", "Не удалось начать запись"))
                self.root.after(0, self._stop_recording_ui)
                
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Ошибка", f"Ошибка записи: {e}"))
            self.root.after(0, self._stop_recording_ui)
    
    def _update_timer(self):
        """Обновление таймера с автоматической остановкой"""
        if hasattr(self, 'start_time') and hasattr(self, 'recording_duration'):
            elapsed = int(time.time() - self.start_time)
            remaining = max(0, self.recording_duration - elapsed)
            
            # Форматируем время
            elapsed_min = elapsed // 60
            elapsed_sec = elapsed % 60
            total_min = self.recording_duration // 60
            total_sec = self.recording_duration % 60
            
            self.timer_label.config(text=f"{elapsed_min:02d}:{elapsed_sec:02d} / {total_min:02d}:{total_sec:02d}")
            
            # Проверяем, не истекло ли время записи
            if elapsed >= self.recording_duration:
                # Автоматически останавливаем запись
                print("⏰ Время записи истекло, останавливаем...")
                self.stop_recording()
                return
            
            # Продолжаем обновление каждую секунду
            self.root.after(1000, self._update_timer)
    
    def stop_recording(self):
        """Остановить запись"""
        try:
            # Останавливаем мониторинг
            self.monitoring_active = False
            
            # Останавливаем запись
            saved_files = self.audio_core.stop_recording()
            
            # Обновляем интерфейс
            self._stop_recording_ui()
            
            # Обновляем список записей
            self.refresh_recordings_list()
            
            # Автоматический анализ если включен
            if self.enable_analysis_var.get() and saved_files:
                outside_path = saved_files.get('outside', {}).get('filepath')
                inside_path = saved_files.get('inside', {}).get('filepath')
                
                if outside_path and inside_path:
                    test_name = self.test_name_var.get()
                    self._analyze_recording(outside_path, inside_path, test_name)
            
            self.status_var.set("✅ Запись завершена")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка остановки записи: {e}")
            self._stop_recording_ui()
    
    def _stop_recording_ui(self):
        """Обновить интерфейс после остановки записи"""
        self.record_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.record_status.config(text="Ожидание", foreground="blue")
        self.timer_label.config(text="00:00 / 00:00")
        
        # Выключаем и скрываем индикаторы
        self.outside_indicator.set_active(False)
        self.inside_indicator.set_active(False)
        self.hide_indicators()
    
    def show_device_summary(self):
        """Показать сводку по устройствам"""
        try:
            devices = self.audio_core.get_audio_devices()
            
            if not devices:
                messagebox.showinfo("Устройства", "Устройства не найдены")
                return
            
            summary = "📊 Сводка по аудиоустройствам:\n\n"
            for i, device in enumerate(devices):
                summary += f"Устройство {i}:\n"
                summary += f"  Название: {device['name']}\n"
                summary += f"  Каналы: {device['channels']}\n"
                summary += f"  Частота: {device.get('sample_rate', 'N/A')} Гц\n"
                summary += "  ---\n"
            
            summary += f"\nВсего устройств: {len(devices)}"
            
            # Создаем отдельное окно для отображения
            summary_window = tk.Toplevel(self.root)
            summary_window.title("Сводка устройств")
            summary_window.geometry("600x400")
            
            text_widget = scrolledtext.ScrolledText(summary_window, wrap=tk.WORD, width=70, height=20)
            text_widget.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            text_widget.insert(tk.END, summary)
            text_widget.config(state=tk.DISABLED)
            
            ttk.Button(summary_window, text="Закрыть", command=summary_window.destroy).pack(pady=5)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка получения сводки: {e}")
    
    def setup_analysis_tab(self, parent):
        """Вкладка анализа"""
        # Заголовок
        ttk.Label(parent, text="АНАЛИЗ ЗАПИСЕЙ", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        # Список записей
        list_frame = ttk.LabelFrame(parent, text="Доступные записи", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # TreeView
        columns = ("name", "date", "duration", "size", "status", "engine")
        self.recordings_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=12)
        
        # Заголовки
        self.recordings_tree.heading("name", text="Имя теста")
        self.recordings_tree.heading("date", text="Дата")
        self.recordings_tree.heading("duration", text="Длительность")
        self.recordings_tree.heading("size", text="Размер")
        self.recordings_tree.heading("status", text="Статус")
        self.recordings_tree.heading("engine", text="Движок")
        
        # Ширина колонок
        self.recordings_tree.column("name", width=180)
        self.recordings_tree.column("date", width=140)
        self.recordings_tree.column("duration", width=80)
        self.recordings_tree.column("size", width=70)
        self.recordings_tree.column("status", width=80)
        self.recordings_tree.column("engine", width=100)
        
        # Прокрутка
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.recordings_tree.yview)
        self.recordings_tree.configure(yscrollcommand=scrollbar.set)
        
        self.recordings_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Двойной клик для анализа
        self.recordings_tree.bind('<Double-1>', lambda e: self.analyze_selected())
        
        # Действия
        action_frame = ttk.Frame(parent)
        action_frame.pack(fill=tk.X, pady=10)
        
        actions = [
            ("📊 Базовый анализ", self.analyze_selected),
            ("🎤 Распознать речь", self.recognize_speech),
            ("🗑️ Удалить запись", self.delete_recording),
            ("📋 Отчет", self.generate_report),
            ("🎵 Воспроизвести", self.play_recording)
        ]
        
        for text, command in actions:
            ttk.Button(action_frame, text=text, command=command).pack(side=tk.LEFT, padx=5)
        
        # Область результатов
        result_frame = ttk.LabelFrame(parent, text="Результаты анализа", padding="10")
        result_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.result_text = scrolledtext.ScrolledText(result_frame, height=10, wrap=tk.WORD)
        self.result_text.pack(fill=tk.BOTH, expand=True)
        self.result_text.config(state=tk.DISABLED)
    
    def setup_engine_tab(self, parent):
        """Вкладка настройки движков распознавания"""
        # Выбор движка
        engine_frame = ttk.LabelFrame(parent, text="Движок распознавания речи (ОФФЛАЙН)", padding="10")
        engine_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(engine_frame, text="Выберите модель:", font=('Arial', 10)).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        # Список доступных движков
        self.engine_combo = ttk.Combobox(engine_frame, width=40, state="readonly")
        
        # Загружаем список движков
        engines = self._get_available_engines()
        self.engine_combo['values'] = engines
        
        if engines:
            self.engine_combo.current(0)
        
        self.engine_combo.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
        
        # Кнопки
        btn_frame = ttk.Frame(engine_frame)
        btn_frame.grid(row=0, column=2, padx=10)
        
        ttk.Button(btn_frame, text="Выбрать", command=self.select_engine).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🔄 Проверить", command=self.check_available_models).pack(side=tk.LEFT, padx=2)
        
        # Статус движка
        self.engine_status_var = tk.StringVar(value="Движок не выбран")
        ttk.Label(engine_frame, textvariable=self.engine_status_var, foreground="blue").grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Кнопка загрузки моделей
        ttk.Button(engine_frame, text="📥 Загрузить модели", 
                  command=self.download_models).grid(row=2, column=0, columnspan=3, pady=10)
        
        # Описание движков
        desc_frame = ttk.LabelFrame(parent, text="Доступные офлайн модели", padding="10")
        desc_frame.pack(fill=tk.X, pady=10)
        
        descriptions = """Для дипломной работы рекомендуется использовать 2 модели:

1. 🎯 Whisper Small (оптимальная) - 500 МБ
   • Международная модель OpenAI
   • Хорошее качество для русского языка
   • Баланс скорости и точности

2. 🇷🇺 Vosk Small RU (русская) - 40 МБ
   • Специализирована для русского языка
   • Очень быстрая
   • Маленький размер

Все модели работают полностью ОФФЛАЙН без API ключей!"""
        
        desc_label = ttk.Label(desc_frame, text=descriptions, justify=tk.LEFT)
        desc_label.pack(anchor=tk.W)
        
        # Тест распознавания
        test_frame = ttk.LabelFrame(parent, text="Тест распознавания", padding="10")
        test_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(test_frame, text="Тестовый аудиофайл:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.test_audio_path = tk.StringVar()
        ttk.Entry(test_frame, textvariable=self.test_audio_path, width=40).grid(row=0, column=1, padx=10, pady=5)
        ttk.Button(test_frame, text="Обзор...", command=self.browse_test_audio).grid(row=0, column=2, padx=5)
        
        ttk.Button(test_frame, text="🧪 Тест распознавания", command=self.test_recognition).grid(row=1, column=0, columnspan=3, pady=10)
        
        # Результаты теста
        self.test_result_text = scrolledtext.ScrolledText(test_frame, height=6, width=60, wrap=tk.WORD)
        self.test_result_text.grid(row=2, column=0, columnspan=3, pady=5, sticky="nsew")
        test_frame.columnconfigure(1, weight=1)
    
    def _get_available_engines(self):
        """Получение списка доступных движков"""
        engines = []
    
        # Проверяем наличие моделей
        # Whisper
        whisper_models = []
        for model in ["tiny", "base", "small", "medium"]:
            if os.path.exists(f"models/whisper/{model}.pt"):
                whisper_models.append(f"whisper-{model}")
    
        # Vosk
        vosk_models = []
        for model in ["small-ru", "large-ru"]:
            if os.path.exists(f"models/vosk/{model}"):
                vosk_models.append(f"vosk-{model}")
    
        # Добавляем все доступные
        engines.extend(whisper_models)
        engines.extend(vosk_models)
    
        # Если нет моделей, показываем инструкцию
        if not engines:
            engines = ["⚠️ Нет моделей. Загрузите модели!"]
    
        return engines
    
    def check_available_models(self):
        """Проверка доступных моделей"""
        available = []
        missing = []
        
        # Проверяем Whisper
        for model in ["tiny", "small", "medium"]:
            path = f"models/whisper/{model}.pt"
            if os.path.exists(path):
                available.append(f"whisper-{model}")
            else:
                missing.append(f"whisper-{model}")
        
        # Проверяем Vosk
        for model in ["small-ru"]:
            path = f"models/vosk/{model}"
            if os.path.exists(path):
                available.append(f"vosk-{model}")
            else:
                missing.append(f"vosk-{model}")
        
        # Показываем результат
        result = "✅ Доступные модели:\n"
        for model in available:
            result += f"  • {model}\n"
        
        if missing:
            result += "\n❌ Отсутствующие модели:\n"
            for model in missing:
                result += f"  • {model}\n"
        
        result += f"\n📊 Всего: {len(available)} из 2 моделей для диплома\n"
        
        if len(available) >= 2:
            result += "🎉 Все модели для диплома готовы!"
        else:
            result += "⚠️ Загрузите недостающие модели через 'Загрузить модели'"
        
        messagebox.showinfo("Проверка моделей", result)
        
        # Обновляем список в комбобоксе
        engines = self._get_available_engines()
        self.engine_combo['values'] = engines
        if engines and "⚠️" not in engines[0]:
            self.engine_combo.current(0)
    
    def setup_export_tab(self, parent):
        """Вкладка экспорта"""
        # Заголовок
        ttk.Label(parent, text="ЭКСПОРТ ДАННЫХ", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        # Блок 1: Экспорт результатов
        export_frame = ttk.LabelFrame(parent, text="Экспорт результатов", padding="10")
        export_frame.pack(fill=tk.X, pady=10)
        
        # Форматы экспорта
        formats_frame = ttk.Frame(export_frame)
        formats_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(formats_frame, text="Формат:").pack(side=tk.LEFT, padx=5)
        
        self.export_format = tk.StringVar(value="csv")
        formats = [("CSV", "csv"), ("JSON", "json"), ("Excel", "excel"), ("Все форматы", "all")]
        
        for text, value in formats:
            ttk.Radiobutton(formats_frame, text=text, value=value, 
                          variable=self.export_format).pack(side=tk.LEFT, padx=10)
        
        # Выбор записей
        selection_frame = ttk.Frame(export_frame)
        selection_frame.pack(fill=tk.X, pady=5)
        
        self.export_selection = tk.StringVar(value="all")
        ttk.Radiobutton(selection_frame, text="Все записи", value="all",
                       variable=self.export_selection).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(selection_frame, text="Только выбранные", value="selected",
                       variable=self.export_selection).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(selection_frame, text="За последние 7 дней", value="week",
                       variable=self.export_selection).pack(side=tk.LEFT, padx=10)
        
        # Кнопка экспорта
        ttk.Button(export_frame, text="📁 Экспортировать данные", 
                  command=self.export_data).pack(pady=10)
        
        # Блок 2: Системная информация
        info_frame = ttk.LabelFrame(parent, text="Системная информация", padding="10")
        info_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.system_info = scrolledtext.ScrolledText(info_frame, height=15, wrap=tk.WORD)
        self.system_info.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Кнопка обновления информации
        ttk.Button(info_frame, text="🔄 Обновить информацию", 
                  command=self.update_system_info).pack(pady=5)
    
    def update_system_info(self):
        """Обновление информации о системе"""
        info = f"🧪 Sound Isolation Tester - Дипломная работа\n"
        info += f"📅 Дата: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        info += f"🐍 Python: {sys.version.split()[0]}\n"
        info += f"💻 ОС: {sys.platform}\n"
        info += f"📁 Папка проекта: {os.path.abspath('.')}\n"
        
        # Подсчет записей
        if os.path.exists(self.recordings_folder):
            wav_files = [f for f in os.listdir(self.recordings_folder) if f.endswith('.wav')]
            info += f"🎙️ Записей: {len(wav_files) // 2}\n"
        else:
            info += f"🎙️ Записей: папка не найдена\n"
    
        # Проверка моделей
        info += "\n🔍 Проверка моделей:\n"
        models_found = 0
    
        # Whisper
        for model in ["tiny", "small", "medium"]:
            if os.path.exists(f"models/whisper/{model}.pt"):
                info += f"  ✅ Whisper {model}\n"
                models_found += 1
            else:
                info += f"  ❌ Whisper {model} (отсутствует)\n"
    
        # Vosk
        if os.path.exists("models/vosk/small-ru"):
            info += f"  ✅ Vosk small-ru\n"
            models_found += 1
        else:
            info += f"  ❌ Vosk small-ru (отсутствует)\n"
    
        info += f"\n📊 Всего моделей: {models_found}/2\n"
    
        if models_found >= 2:
            info += "✅ Все модели для диплома готовы!"
        else:
            info += "⚠️ Загрузите недостающие модели!"
    
        self.system_info.delete(1.0, tk.END)
        self.system_info.insert(1.0, info)
        self.system_info.config(state=tk.DISABLED)
    
    def _analyze_recording(self, outside_path, inside_path, test_name):
        """Анализ записи"""
        try:
            self.status_var.set("📊 Анализ записи...")
            
            analysis = self.analyzer.analyze_with_audio_analysis(
                outside_path, inside_path, test_name,
                enable_speech_recognition=bool(self.recognizer)
            )
            
            # Показываем результаты
            self._display_analysis_results(analysis)
            
            self.status_var.set("✅ Анализ завершен")
            
        except Exception as e:
            self.status_var.set("❌ Ошибка анализа")
            messagebox.showwarning("Предупреждение", f"Ошибка анализа: {e}")
    
    def _display_analysis_results(self, analysis):
        """Отобразить результаты анализа"""
        try:
            overall = analysis.get('results', {}).get('overall_assessment', {})
            
            result_text = "=" * 50 + "\n"
            result_text += f"АНАЛИЗ ЗАПИСИ: {analysis.get('test_name', 'N/A')}\n"
            result_text += f"ВРЕМЯ: {analysis.get('timestamp', 'N/A')}\n"
            result_text += "=" * 50 + "\n\n"
            
            # Вердикт
            verdict = overall.get('verdict', 'N/A')
            color = overall.get('color', 'black')
            result_text += f"ВЕРДИКТ: {verdict}\n\n"
            
            # Сводка
            summary = overall.get('summary', 'N/A')
            result_text += f"СВОДКА: {summary}\n\n"
            
            # Детальные метрики
            detailed = analysis.get('results', {}).get('detailed_metrics', {})
            if detailed:
                basic = detailed.get('basic', {})
                if basic:
                    result_text += "ОСНОВНЫЕ МЕТРИКИ:\n"
                    result_text += f"  • Ослабление: {basic.get('attenuation_db', 0):.1f} дБ\n"
                    result_text += f"  • Качество изоляции: {basic.get('isolation_quality', 'N/A')}\n"
                    result_text += f"  • Корреляция сигналов: {basic.get('correlation', 0):.3f}\n"
                
                composite = detailed.get('composite_scores', {})
                if composite:
                    result_text += "\nКОМПОЗИТНЫЕ ОЦЕНКИ:\n"
                    result_text += f"  • Общая оценка: {composite.get('total_score', 0):.1f}/100\n"
                    result_text += f"  • Оценка: {composite.get('grade', 'N/A')}\n"
            
            # Рекомендации
            recommendations = overall.get('recommendations', [])
            if recommendations:
                result_text += "\nРЕКОМЕНДАЦИИ:\n"
                for i, rec in enumerate(recommendations, 1):
                    result_text += f"  {i}. {rec}\n"
            
            result_text += "\n" + "=" * 50
            
            # Отображаем в интерфейсе
            self.result_text.config(state=tk.NORMAL)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result_text)
            
            # Настраиваем цвет вердикта
            self.result_text.tag_add("verdict", "3.0", "3.end")
            self.result_text.tag_config("verdict", foreground=color, font=('Arial', 10, 'bold'))
            
            self.result_text.config(state=tk.DISABLED)
            
        except Exception as e:
            print(f"Ошибка отображения результатов: {e}")
    
    def refresh_recordings_list(self):
        """Обновить список записей"""
        try:
            # Очищаем дерево
            for item in self.recordings_tree.get_children():
                self.recordings_tree.delete(item)
            
            # Получаем список записей
            recordings = self._get_recordings_list()
            
            # Добавляем записи в дерево
            for rec in recordings:
                self.recordings_tree.insert("", tk.END, values=(
                    rec.get('test_name', 'N/A'),
                    rec.get('timestamp', 'N/A'),
                    rec.get('duration', 'N/A'),
                    rec.get('size', 'N/A'),
                    rec.get('status', 'N/A'),
                    rec.get('engine', 'N/A')
                ))
            
        except Exception as e:
            print(f"Ошибка обновления списка записей: {e}")
    
    def _get_recordings_list(self):
        """Получить список записей"""
        recordings = []
        
        try:
            # Сканируем папку recordings
            for file in os.listdir(self.recordings_folder):
                if file.endswith('_metadata.json'):
                    metadata_path = os.path.join(self.recordings_folder, file)
                    try:
                        with open(metadata_path, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                            
                            # Формируем информацию о записи
                            rec_info = {
                                'test_name': metadata.get('test_name', file.replace('_metadata.json', '')),
                                'timestamp': metadata.get('timestamp', 'N/A'),
                                'duration': f"{metadata.get('duration', 0):.1f} сек",
                                'size': self._get_recording_size(metadata),
                                'status': '✅' if metadata.get('analysis_ready', False) else '⚠️',
                                'engine': 'N/A'
                            }
                            recordings.append(rec_info)
                            
                    except Exception as e:
                        print(f"Ошибка чтения {file}: {e}")
            
            # Сортируем по дате (сначала новые)
            recordings.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
            
        except Exception as e:
            print(f"Ошибка получения списка записей: {e}")
        
        return recordings
    
    def _get_recording_size(self, metadata):
        """Получить размер записи"""
        try:
            files = metadata.get('files', {})
            total_size = 0
            
            for channel in ['outside', 'inside']:
                file_info = files.get(channel, {})
                filepath = file_info.get('filepath')
                if filepath and os.path.exists(filepath):
                    total_size += os.path.getsize(filepath)
            
            # Конвертируем в КБ/МБ
            if total_size > 1024 * 1024:
                return f"{total_size / (1024 * 1024):.1f} МБ"
            else:
                return f"{total_size / 1024:.0f} КБ"
                
        except:
            return "N/A"
    
    def analyze_selected(self):
        """Анализ выбранной записи"""
        try:
            selection = self.recordings_tree.selection()
            if not selection:
                messagebox.showwarning("Предупреждение", "Выберите запись для анализа")
                return
            
            # Получаем данные выбранной записи
            item = self.recordings_tree.item(selection[0])
            test_name = item['values'][0]
            
            # Находим файлы записи
            outside_path = os.path.join(self.recordings_folder, f"{test_name}_outside.wav")
            inside_path = os.path.join(self.recordings_folder, f"{test_name}_inside.wav")
            
            if not os.path.exists(outside_path) or not os.path.exists(inside_path):
                messagebox.showerror("Ошибка", "Файлы записи не найдены")
                return
            
            # Выполняем анализ
            self.status_var.set("📊 Анализ записи...")
            
            analysis = self.analyzer.analyze_with_audio_analysis(
                outside_path, inside_path, test_name,
                enable_speech_recognition=bool(self.recognizer)
            )
            
            # Отображаем результаты
            self._display_analysis_results(analysis)
            
            self.status_var.set("✅ Анализ завершен")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка анализа: {e}")
    
    def recognize_speech(self):
        """Распознавание речи в выбранной записи"""
        try:
            if not self.recognizer:
                messagebox.showwarning("Предупреждение", 
                    "Модуль распознавания речи недоступен.\n"
                    "1. Установите модели через вкладку 'Движки'\n"
                    "2. Выберите движок распознавания")
                return
            
            selection = self.recordings_tree.selection()
            if not selection:
                messagebox.showwarning("Предупреждение", "Выберите запись для распознавания")
                return
            
            # Получаем данные выбранной записи
            item = self.recordings_tree.item(selection[0])
            test_name = item['values'][0]
            
            # Находим файлы записи
            outside_path = os.path.join(self.recordings_folder, f"{test_name}_outside.wav")
            inside_path = os.path.join(self.recordings_folder, f"{test_name}_inside.wav")
            
            if not os.path.exists(outside_path) or not os.path.exists(inside_path):
                messagebox.showerror("Ошибка", "Файлы записи не найдены")
                return
            
            # Распознавание речи
            self.status_var.set("🎤 Распознавание речи...")
            
            result = self.recognizer.analyze_pair(outside_path, inside_path)
            
            # Отображаем результаты
            result_text = "=" * 50 + "\n"
            result_text += f"РАСПОЗНАВАНИЕ РЕЧИ: {test_name}\n"
            result_text += f"ДВИЖОК: {result.get('engine', 'N/A')}\n"
            result_text += "=" * 50 + "\n\n"
            
            # Текст снаружи
            outside = result.get('outside', {})
            result_text += "СНАРУЖИ:\n"
            result_text += f"  Текст: {outside.get('text', 'N/A')}\n"
            result_text += f"  Уверенность: {outside.get('confidence', 0):.2f}\n"
            result_text += f"  Слов: {outside.get('word_count', 0)}\n\n"
            
            # Текст внутри
            inside = result.get('inside', {})
            result_text += "ВНУТРИ:\n"
            result_text += f"  Текст: {inside.get('text', 'N/A')}\n"
            result_text += f"  Уверенность: {inside.get('confidence', 0):.2f}\n"
            result_text += f"  Слов: {inside.get('word_count', 0)}\n\n"
            
            # Сравнение
            comparison = result.get('comparison', {})
            result_text += "СРАВНЕНИЕ:\n"
            result_text += f"  WER (ошибок на слово): {comparison.get('wer', 0):.2%}\n"
            
            if comparison.get('leakage_detected', False):
                result_text += f"  ⚠️ ОБНАРУЖЕНА УТЕЧКА РЕЧИ!\n"
                result_text += f"  Уровень утечки: {comparison.get('leakage_score', 0):.2f}\n"
            else:
                result_text += f"  ✅ Утечка не обнаружена\n"
            
            result_text += f"  Время обработки: {comparison.get('total_processing_time', 0):.1f} сек\n"
            
            result_text += "\n" + "=" * 50
            
            # Отображаем в интерфейсе
            self.result_text.config(state=tk.NORMAL)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result_text)
            self.result_text.config(state=tk.DISABLED)
            
            self.status_var.set("✅ Распознавание завершено")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка распознавания: {e}")
    
    def delete_recording(self):
        """Удалить выбранную запись"""
        try:
            selection = self.recordings_tree.selection()
            if not selection:
                messagebox.showwarning("Предупреждение", "Выберите запись для удаления")
                return
            
            # Подтверждение
            item = self.recordings_tree.item(selection[0])
            test_name = item['values'][0]
            
            confirm = messagebox.askyesno(
                "Подтверждение удаления",
                f"Вы уверены, что хотите удалить запись '{test_name}'?\n"
                f"Все связанные файлы будут удалены безвозвратно."
            )
            
            if not confirm:
                return
            
            # Удаляем файлы
            files_to_delete = [
                os.path.join(self.recordings_folder, f"{test_name}_outside.wav"),
                os.path.join(self.recordings_folder, f"{test_name}_inside.wav"),
                os.path.join(self.recordings_folder, f"{test_name}_metadata.json")
            ]
            
            deleted_count = 0
            for filepath in files_to_delete:
                if os.path.exists(filepath):
                    try:
                        os.remove(filepath)
                        deleted_count += 1
                    except Exception as e:
                        print(f"Ошибка удаления {filepath}: {e}")
            
            # Обновляем список
            self.refresh_recordings_list()
            
            messagebox.showinfo("Успех", f"Удалено записей: {deleted_count}/3")
            self.status_var.set("🗑️ Запись удалена")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка удаления: {e}")
    
    def generate_report(self):
        """Сгенерировать отчет по выбранной записи"""
        try:
            selection = self.recordings_tree.selection()
            if not selection:
                messagebox.showwarning("Предупреждение", "Выберите запись для отчета")
                return
            
            # Получаем данные выбранной записи
            item = self.recordings_tree.item(selection[0])
            test_name = item['values'][0]
            
            # Создаем диалог выбора формата
            format_window = tk.Toplevel(self.root)
            format_window.title("Выбор формата отчета")
            format_window.geometry("400x200")
            format_window.transient(self.root)
            format_window.grab_set()
            
            # Центрируем
            format_window.update_idletasks()
            x = (self.root.winfo_screenwidth() // 2) - (400 // 2)
            y = (self.root.winfo_screenheight() // 2) - (200 // 2)
            format_window.geometry(f'400x200+{x}+{y}')
            
            ttk.Label(format_window, text="📄 ВЫБЕРИТЕ ФОРМАТ ОТЧЕТА", 
                     font=('Arial', 12, 'bold')).pack(pady=10)
            
            format_var = tk.StringVar(value="html")
            
            def create_report(format_type):
                format_window.destroy()
                self._create_report_file(test_name, format_type)
            
            ttk.Radiobutton(format_window, text="📄 HTML (для печати)", 
                           value="html", variable=format_var).pack(pady=5)
            ttk.Radiobutton(format_window, text="📝 Текстовый (TXT)", 
                           value="txt", variable=format_var).pack(pady=5)
            ttk.Radiobutton(format_window, text="📊 Excel (XLSX)", 
                           value="excel", variable=format_var).pack(pady=5)
            
            ttk.Button(format_window, text="Создать отчет", 
                      command=lambda: create_report(format_var.get())).pack(pady=15)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка генерации отчета: {e}")
    
    def _create_report_file(self, test_name, format_type):
        """Создать файл отчета в указанном формате"""
        try:
            # Находим метаданные
            metadata_path = os.path.join(self.recordings_folder, f"{test_name}_metadata.json")
            if not os.path.exists(metadata_path):
                messagebox.showerror("Ошибка", "Метаданные записи не найдены")
                return
            
            # Читаем метаданные
            with open(metadata_path, 'r', encoding='utf-8') as f:
                metadata = json.load(f)
            
            # Ищем файлы анализа
            analysis_path = os.path.join(self.recordings_folder, f"{test_name}_analysis.json")
            analysis_data = None
            if os.path.exists(analysis_path):
                with open(analysis_path, 'r', encoding='utf-8') as f:
                    analysis_data = json.load(f)
            
            # Запрашиваем место сохранения
            if format_type == "html":
                ext = ".html"
                filetypes = [("HTML файлы", "*.html"), ("Все файлы", "*.*")]
                initialfile = f"{test_name}_report.html"
            elif format_type == "excel":
                ext = ".xlsx"
                filetypes = [("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
                initialfile = f"{test_name}_report.xlsx"
            else:
                ext = ".txt"
                filetypes = [("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")]
                initialfile = f"{test_name}_report.txt"
            
            filename = filedialog.asksaveasfilename(
                defaultextension=ext,
                filetypes=filetypes,
                initialfile=initialfile
            )
            
            if not filename:
                return
            
            # Создаем отчет в выбранном формате
            if format_type == "html":
                self._create_html_report(metadata, analysis_data, filename)
            elif format_type == "excel":
                self._create_excel_report(metadata, analysis_data, filename)
            else:
                self._create_text_report(metadata, analysis_data, filename)
            
            # Показываем сообщение об успехе с опцией открытия
            if format_type == "html":
                open_result = messagebox.askyesno("Успех", 
                    f"HTML-отчет сохранен:\n{filename}\n\n"
                    f"Открыть в браузере для печати?")
                
                if open_result:
                    webbrowser.open('file://' + os.path.abspath(filename))
            else:
                messagebox.showinfo("Успех", f"Отчет сохранен:\n{filename}")
            
            self.status_var.set("📋 Отчет сгенерирован")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка создания отчета: {e}")
    
    def _create_html_report(self, metadata, analysis_data, filename):
        """Создать HTML отчет для печати"""
        
        # Получаем данные из метаданных
        test_name = metadata.get('test_name', 'Неизвестный тест')
        timestamp = metadata.get('timestamp', 'Нет данных')
        duration = metadata.get('duration', 0)
        sample_rate = metadata.get('sample_rate', 0)
        
        # Получаем результаты анализа если есть
        overall_score = "Н/Д"
        if analysis_data:
            results = analysis_data.get('results', {})
            overall = results.get('overall_assessment', {})
            overall_score = overall.get('verdict', 'Н/Д')
            grade = overall.get('grade', 'Н/Д')
            color = overall.get('color', 'black')
            recommendations = overall.get('recommendations', [])
        
        # Создаем HTML документ
        html_content = f'''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Отчет по тесту звукоизоляции - {test_name}</title>
    <style>
        @media print {{
            @page {{
                margin: 2cm;
                size: A4;
            }}
            body {{
                font-size: 12pt;
            }}
            .page-break {{
                page-break-before: always;
            }}
            .no-print {{
                display: none;
            }}
        }}
        
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        
        body {{
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 210mm;
            margin: 0 auto;
            padding: 20mm;
            background-color: #f9f9f9;
        }}
        
        .header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 3px solid #2c3e50;
        }}
        
        .header h1 {{
            color: #2c3e50;
            font-size: 24pt;
            margin-bottom: 10px;
        }}
        
        .header .subtitle {{
            color: #7f8c8d;
            font-size: 14pt;
        }}
        
        .info-card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border-left: 5px solid #3498db;
        }}
        
        .result-card {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border-left: 5px solid #2ecc71;
        }}
        
        .verdict-card {{
            background: white;
            border-radius: 8px;
            padding: 30px;
            margin: 30px 0;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            text-align: center;
            border: 2px solid #e74c3c;
        }}
        
        .verdict-card h2 {{
            color: #e74c3c;
            font-size: 20pt;
            margin-bottom: 15px;
        }}
        
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        
        .metric-item {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 6px;
            text-align: center;
            border: 1px solid #dee2e6;
        }}
        
        .metric-value {{
            font-size: 24pt;
            font-weight: bold;
            color: #2c3e50;
            margin: 10px 0;
        }}
        
        .metric-label {{
            color: #6c757d;
            font-size: 11pt;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        
        h2 {{
            color: #2c3e50;
            margin: 25px 0 15px 0;
            padding-bottom: 10px;
            border-bottom: 2px solid #ecf0f1;
            font-size: 18pt;
        }}
        
        h3 {{
            color: #34495e;
            margin: 20px 0 10px 0;
            font-size: 14pt;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
            font-size: 11pt;
        }}
        
        table th {{
            background: #2c3e50;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        
        table td {{
            padding: 12px;
            border-bottom: 1px solid #ddd;
        }}
        
        table tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        
        .recommendations {{
            background: #fff3cd;
            border-left: 5px solid #ffc107;
            padding: 20px;
            margin: 20px 0;
            border-radius: 6px;
        }}
        
        .recommendations ul {{
            padding-left: 20px;
            margin: 10px 0;
        }}
        
        .recommendations li {{
            margin: 8px 0;
        }}
        
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #ecf0f1;
            text-align: center;
            color: #7f8c8d;
            font-size: 10pt;
        }}
        
        .print-button {{
            display: block;
            width: 200px;
            margin: 30px auto;
            padding: 12px 24px;
            background: #3498db;
            color: white;
            text-align: center;
            text-decoration: none;
            border-radius: 6px;
            font-weight: bold;
            cursor: pointer;
            border: none;
            font-size: 12pt;
        }}
        
        .print-button:hover {{
            background: #2980b9;
        }}
        
        .badge {{
            display: inline-block;
            padding: 5px 10px;
            border-radius: 20px;
            font-size: 10pt;
            font-weight: bold;
            margin: 0 5px;
        }}
        
        .badge-success {{
            background: #d4edda;
            color: #155724;
        }}
        
        .badge-warning {{
            background: #fff3cd;
            color: #856404;
        }}
        
        .badge-danger {{
            background: #f8d7da;
            color: #721c24;
        }}
        
        .grade {{
            font-size: 32pt;
            font-weight: bold;
            color: #2c3e50;
            text-align: center;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 ОТЧЕТ ПО ТЕСТУ ЗВУКОИЗОЛЯЦИИ</h1>
        <div class="subtitle">Дипломная работа - Sound Isolation Tester v3.13</div>
    </div>
    
    <div class="info-card">
        <h2>📋 ИНФОРМАЦИЯ О ТЕСТЕ</h2>
        <div class="metrics-grid">
            <div class="metric-item">
                <div class="metric-label">Название теста</div>
                <div class="metric-value">{test_name}</div>
            </div>
            <div class="metric-item">
                <div class="metric-label">Дата и время</div>
                <div class="metric-value">{timestamp}</div>
            </div>
            <div class="metric-item">
                <div class="metric-label">Длительность</div>
                <div class="metric-value">{duration:.1f} сек</div>
            </div>
            <div class="metric-item">
                <div class="metric-label">Частота дискретизации</div>
                <div class="metric-value">{sample_rate} Гц</div>
            </div>
        </div>
    </div>
    
    <div class="result-card">
        <h2>📈 РЕЗУЛЬТАТЫ АНАЛИЗА</h2>
        '''
        
        # Добавляем результаты анализа если есть
        if analysis_data:
            results = analysis_data.get('results', {})
            overall = results.get('overall_assessment', {})
            detailed = results.get('detailed_metrics', {})
            
            html_content += f'''
            <div class="verdict-card">
                <h2>ВЕРДИКТ</h2>
                <div class="grade">{overall.get('verdict', 'Н/Д')}</div>
                <p style="font-size: 14pt; margin-top: 10px;">{overall.get('summary', 'Нет данных')}</p>
            </div>
            
            <h3>Детальные метрики</h3>
            '''
            
            # Базовые метрики
            if detailed.get('basic'):
                basic = detailed['basic']
                html_content += f'''
                <table>
                    <tr>
                        <th>Параметр</th>
                        <th>Значение</th>
                        <th>Оценка</th>
                    </tr>
                    <tr>
                        <td>Ослабление звука</td>
                        <td>{basic.get('attenuation_db', 0):.1f} дБ</td>
                        <td>{basic.get('attenuation_rating', 'Н/Д')}</td>
                    </tr>
                    <tr>
                        <td>Качество изоляции</td>
                        <td>{basic.get('isolation_quality', 'Н/Д')}</td>
                        <td>{basic.get('isolation_rating', 'Н/Д')}</td>
                    </tr>
                    <tr>
                        <td>Корреляция сигналов</td>
                        <td>{basic.get('correlation', 0):.3f}</td>
                        <td>{basic.get('correlation_rating', 'Н/Д')}</td>
                    </tr>
                </table>
                '''
            
            # Композитные оценки
            if detailed.get('composite_scores'):
                composite = detailed['composite_scores']
                html_content += f'''
                <h3>Композитные оценки</h3>
                <div class="metrics-grid">
                    <div class="metric-item">
                        <div class="metric-label">Общая оценка</div>
                        <div class="metric-value">{composite.get('total_score', 0):.1f}/100</div>
                    </div>
                    <div class="metric-item">
                        <div class="metric-label">Оценка по шкале</div>
                        <div class="metric-value">{composite.get('grade', 'Н/Д')}</div>
                    </div>
                    <div class="metric-item">
                        <div class="metric-label">Эффективность</div>
                        <div class="metric-value">{composite.get('effectiveness_percent', 0):.1f}%</div>
                    </div>
                </div>
                '''
            
            # Рекомендации
            recommendations = overall.get('recommendations', [])
            if recommendations:
                html_content += '''
                <div class="recommendations">
                    <h3>🏆 РЕКОМЕНДАЦИИ ПО УЛУЧШЕНИЮ</h3>
                    <ul>
                '''
                for rec in recommendations:
                    html_content += f'<li>{rec}</li>'
                html_content += '</ul></div>'
        
        # Если анализа нет, показываем информационное сообщение
        else:
            html_content += '''
            <div style="text-align: center; padding: 40px; background: #f8f9fa; border-radius: 8px;">
                <h3>⚠️ Анализ не выполнен</h3>
                <p>Для данной записи не выполнен анализ звукоизоляции.</p>
                <p>Выполните анализ через вкладку "АНАЛИЗ" для получения подробных результатов.</p>
            </div>
            '''
        
        # Технические данные
        html_content += f'''
        </div>
        
        <div class="info-card">
            <h2>🔧 ТЕХНИЧЕСКИЕ ДАННЫЕ</h2>
            <table>
                <tr>
                    <th>Параметр</th>
                    <th>Значение</th>
                </tr>
        '''
        
        # Добавляем технические данные из metadata
        if 'files' in metadata:
            files = metadata['files']
            for channel, data in files.items():
                html_content += f'''
                <tr>
                    <td>Файл ({channel})</td>
                    <td>{data.get('filename', 'N/A')}</td>
                </tr>
                <tr>
                    <td>Размер ({channel})</td>
                    <td>{data.get('filesize_mb', 0):.2f} МБ</td>
                </tr>
                <tr>
                    <td>Сэмплов ({channel})</td>
                    <td>{data.get('samples', 0):,}</td>
                </tr>
                '''
        
        html_content += f'''
            </table>
        </div>
        
        <div class="info-card">
            <h2>📊 СИСТЕМНАЯ ИНФОРМАЦИЯ</h2>
            <table>
                <tr>
                    <td>Дата создания отчета</td>
                    <td>{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</td>
                </tr>
                <tr>
                    <td>Версия приложения</td>
                    <td>Sound Isolation Tester v3.13</td>
                </tr>
                <tr>
                    <td>Операционная система</td>
                    <td>{sys.platform}</td>
                </tr>
                <tr>
                    <td>Версия Python</td>
                    <td>{sys.version.split()[0]}</td>
                </tr>
            </table>
        </div>
        
        <div class="footer">
            <p>© {datetime.now().year} - Дипломная работа "Разработка системы тестирования звукоизоляции"</p>
            <p>Отчет сгенерирован автоматически. Для печати нажмите Ctrl+P</p>
            <p>Все данные конфиденциальны и предназначены только для академического использования</p>
        </div>
        
        <button class="print-button no-print" onclick="window.print()">🖨️ Печать отчета</button>
        
        <script>
            // Автоматически предлагаем печать после загрузки
            window.onload = function() {{
                // Автоматическое открытие диалога печати (можно закомментировать если не нужно)
                // setTimeout(() => {{ window.print(); }}, 1000);
            }};
        </script>
    </body>
    </html>
        '''
        
        # Сохраняем HTML файл
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    def _create_excel_report(self, metadata, analysis_data, filename):
        """Создать Excel отчет"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Создаем рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет звукоизоляции"
            
            # Стили
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            title_font = Font(name='Arial', size=14, bold=True, color='2C3E50')
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Заголовок
            ws.merge_cells('A1:F1')
            ws['A1'] = f'ОТЧЕТ ПО ТЕСТУ ЗВУКОИЗОЛЯЦИИ - {metadata.get("test_name", "N/A")}'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align
            
            # Основная информация
            ws['A3'] = 'Основная информация'
            ws['A3'].font = Font(bold=True)
            
            data = [
                ['Параметр', 'Значение'],
                ['Название теста', metadata.get('test_name', 'N/A')],
                ['Дата и время', metadata.get('timestamp', 'N/A')],
                ['Длительность', f"{metadata.get('duration', 0):.1f} сек"],
                ['Частота дискретизации', f"{metadata.get('sample_rate', 0)} Гц"],
            ]
            
            for i, row in enumerate(data, start=3):
                for j, value in enumerate(row, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = value
                    cell.border = border
            
            # Если есть результаты анализа
            if analysis_data:
                results = analysis_data.get('results', {})
                overall = results.get('overall_assessment', {})
                
                # Вердикт
                ws['A8'] = 'Результаты анализа'
                ws['A8'].font = Font(bold=True)
                
                verdict_data = [
                    ['Вердикт', overall.get('verdict', 'Н/Д')],
                    ['Оценка', overall.get('grade', 'Н/Д')],
                    ['Общая оценка', f"{results.get('detailed_metrics', {}).get('composite_scores', {}).get('total_score', 0):.1f}/100"],
                ]
                
                for i, row in enumerate(verdict_data, start=9):
                    for j, value in enumerate(row, start=1):
                        cell = ws.cell(row=i, column=j)
                        cell.value = value
                        cell.border = border
            
            # Настраиваем ширину колонок
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Сохраняем файл
            wb.save(filename)
            
        except ImportError:
            messagebox.showerror("Ошибка", 
                "Для создания Excel отчета установите:\n"
                "pip install pandas openpyxl")
            raise
    
    def _create_text_report(self, metadata, analysis_data, filename):
        """Создать текстовый отчет"""
        report = "=" * 60 + "\n"
        report += "ОТЧЕТ О ТЕСТЕ ЗВУКОИЗОЛЯЦИИ\n"
        report += "=" * 60 + "\n\n"
        
        report += f"Имя теста: {metadata.get('test_name', 'N/A')}\n"
        report += f"Дата и время: {metadata.get('timestamp', 'N/A')}\n"
        report += f"Частота дискретизации: {metadata.get('sample_rate', 'N/A')} Гц\n"
        report += f"Длительность: {metadata.get('duration', 0):.2f} сек\n\n"
        
        if analysis_data:
            results = analysis_data.get('results', {})
            overall = results.get('overall_assessment', {})
            
            report += "РЕЗУЛЬТАТЫ АНАЛИЗА:\n"
            report += "-" * 40 + "\n"
            report += f"Вердикт: {overall.get('verdict', 'Н/Д')}\n"
            report += f"Оценка: {overall.get('grade', 'Н/Д')}\n"
            report += f"Сводка: {overall.get('summary', 'Н/Д')}\n\n"
            
            # Рекомендации
            recommendations = overall.get('recommendations', [])
            if recommendations:
                report += "РЕКОМЕНДАЦИИ:\n"
                report += "-" * 40 + "\n"
                for i, rec in enumerate(recommendations, 1):
                    report += f"{i}. {rec}\n"
                report += "\n"
        
        report += "ТЕХНИЧЕСКИЕ ДАННЫЕ:\n"
        report += "-" * 40 + "\n"
        if 'files' in metadata:
            files = metadata['files']
            for channel, data in files.items():
                report += f"{channel.upper()}:\n"
                report += f"  Файл: {data.get('filename', 'N/A')}\n"
                report += f"  Размер: {data.get('filesize_mb', 0):.2f} МБ\n"
                report += f"  Сэмплов: {data.get('samples', 0):,}\n"
        
        report += "\n" + "=" * 60 + "\n"
        report += "СИСТЕМНАЯ ИНФОРМАЦИЯ:\n"
        report += "-" * 40 + "\n"
        report += f"Дата создания отчета: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report += f"Версия приложения: Sound Isolation Tester v3.13\n"
        report += f"Операционная система: {sys.platform}\n"
        report += f"Версия Python: {sys.version.split()[0]}\n"
        
        report += "\n" + "=" * 60 + "\n"
        report += "ПРИМЕЧАНИЕ:\n"
        report += "-" * 40 + "\n"
        report += "Для более наглядного представления данных рекомендуется\n"
        report += "создать HTML отчет с возможностью печати.\n"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(report)
    
    def play_recording(self):
        """Воспроизвести выбранную запись"""
        try:
            selection = self.recordings_tree.selection()
            if not selection:
                messagebox.showwarning("Предупреждение", "Выберите запись для воспроизведения")
                return
            
            # Получаем данные выбранной записи
            item = self.recordings_tree.item(selection[0])
            test_name = item['values'][0]
            
            # Находим файлы записи
            outside_path = os.path.join(self.recordings_folder, f"{test_name}_outside.wav")
            inside_path = os.path.join(self.recordings_folder, f"{test_name}_inside.wav")
            
            if not os.path.exists(outside_path) or not os.path.exists(inside_path):
                messagebox.showerror("Ошибка", "Файлы записи не найдены")
                return
            
            # Запрашиваем какой канал воспроизводить
            channel = messagebox.askquestion(
                "Выбор канала",
                "Какой канал воспроизвести?",
                detail="'Да' - Снаружи\n'Нет' - Внутри\n'Отмена' - Оба",
                type=messagebox.YESNOCANCEL
            )
            
            if channel == messagebox.YES:
                file_to_play = outside_path
                channel_name = "СНАРУЖИ"
            elif channel == messagebox.NO:
                file_to_play = inside_path
                channel_name = "ВНУТРИ"
            else:
                # Воспроизводим оба
                file_to_play = None
            
            if file_to_play:
                # Воспроизводим один файл
                self._play_audio_file(file_to_play, channel_name)
            else:
                # Воспроизводим оба файла (последовательно)
                self._play_audio_file(outside_path, "СНАРУЖИ")
                time.sleep(1)
                self._play_audio_file(inside_path, "ВНУТРИ")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка воспроизведения: {e}")
    
    def _play_audio_file(self, filepath, channel_name):
        """Воспроизвести аудиофайл"""
        try:
            # Проверяем платформу
            if sys.platform == "win32":
                # Windows
                os.startfile(filepath)
            elif sys.platform == "darwin":
                # macOS
                subprocess.call(["open", filepath])
            else:
                # Linux
                subprocess.call(["xdg-open", filepath])
            
            self.status_var.set(f"🎵 Воспроизведение: {channel_name}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось воспроизвести файл: {e}")
    
    def select_engine(self):
        """Выбрать движок распознавания"""
        try:
            engine_name = self.engine_combo.get()
            
            if not engine_name or "⚠️" in engine_name:
                messagebox.showwarning("Предупреждение", 
                    "Нет доступных моделей!\n"
                    "Загрузите модели через кнопку 'Загрузить модели'")
                return
            
            if self.recognizer:
                # Преобразуем строку в Enum
                try:
                    engine = RecognitionEngine(engine_name)
                except ValueError:
                    # Если движок не в Enum, пробуем создать из строки
                    if engine_name.startswith("whisper-"):
                        engine = RecognitionEngine(engine_name)
                    elif engine_name.startswith("vosk-"):
                        engine = RecognitionEngine(engine_name)
                    else:
                        raise ValueError(f"Неизвестный движок: {engine_name}")
                
                # Устанавливаем движок
                success = self.recognizer.set_engine(engine)
                
                if success:
                    self.current_engine = engine
                    self.engine_status_var.set(f"✅ Выбран: {engine_name}")
                    self.status_var.set(f"Движок установлен: {engine_name}")
                    
                    # Обновляем анализатор
                    self.analyzer.set_recognition_engine(engine_name)
                else:
                    self.engine_status_var.set(f"❌ Ошибка загрузки: {engine_name}")
                    messagebox.showerror("Ошибка", f"Не удалось загрузить движок: {engine_name}")
            else:
                messagebox.showwarning("Предупреждение", 
                    "Модуль распознавания недоступен.\n"
                    "Установите зависимости: pip install vosk whisper")
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка выбора движка: {e}")
    
    def download_models(self):
        """Загрузить модели распознавания"""
        try:
            # Показываем информационное окно
            info_window = tk.Toplevel(self.root)
            info_window.title("Загрузка моделей")
            info_window.geometry("500x300")
            info_window.transient(self.root)
            info_window.grab_set()
            
            # Центрируем
            info_window.update_idletasks()
            x = (self.root.winfo_screenwidth() // 2) - (500 // 2)
            y = (self.root.winfo_screenheight() // 2) - (300 // 2)
            info_window.geometry(f'500x300+{x}+{y}')
            
            # Контент
            ttk.Label(info_window, text="📥 ЗАГРУЗКА МОДЕЛЕЙ", 
                     font=('Arial', 12, 'bold')).pack(pady=10)
            
            info_text = """Для загрузки моделей выполните:

1. Запустите скрипт download_models.py:
   • Откройте командную строку/терминал
   • Перейдите в папку с проектом
   • Выполните: python download_models.py

2. Или выполните вручную:
   • pip install vosk whisper
   • Загрузите модели Whisper:
     https://github.com/openai/whisper
   • Загрузите модели Vosk:
     https://alphacephei.com/vosk/models"""
            
            text_widget = scrolledtext.ScrolledText(info_window, wrap=tk.WORD, 
                                                   width=60, height=12)
            text_widget.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
            text_widget.insert(tk.END, info_text)
            text_widget.config(state=tk.DISABLED)
            
            def open_download_script():
                try:
                    if os.path.exists("download_models.py"):
                        subprocess.Popen([sys.executable, "download_models.py"])
                    else:
                        messagebox.showwarning("Предупреждение", 
                            "Файл download_models.py не найден\n"
                            "Скачайте его из архива проекта")
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось запустить скрипт: {e}")
            
            # Кнопки
            btn_frame = ttk.Frame(info_window)
            btn_frame.pack(pady=10)
            
            ttk.Button(btn_frame, text="🚀 Запустить скрипт", 
                      command=open_download_script).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="❌ Закрыть", 
                      command=info_window.destroy).pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка: {e}")
    
    def browse_test_audio(self):
        """Выбрать тестовый аудиофайл"""
        filename = filedialog.askopenfilename(
            title="Выберите аудиофайл",
            filetypes=[("WAV файлы", "*.wav"), ("Все файлы", "*.*")]
        )
        
        if filename:
            self.test_audio_path.set(filename)
    
    def test_recognition(self):
        """Тест распознавания речи"""
        try:
            audio_path = self.test_audio_path.get()
            
            if not audio_path or not os.path.exists(audio_path):
                messagebox.showwarning("Предупреждение", "Выберите аудиофайл для теста")
                return
            
            if not self.recognizer or not self.current_engine:
                messagebox.showwarning("Предупреждение", 
                    "Сначала выберите движок распознавания")
                return
            
            # Выполняем распознавание
            self.status_var.set("🧪 Тест распознавания...")
            
            result = self.recognizer.transcribe(audio_path)
            
            # Отображаем результаты
            self.test_result_text.delete(1.0, tk.END)
            
            if result and result.text:
                result_text = f"✅ РАСПОЗНАНО УСПЕШНО\n\n"
                result_text += f"Движок: {result.engine}\n"
                result_text += f"Текст: {result.text}\n"
                result_text += f"Уверенность: {result.confidence:.2f}\n"
                result_text += f"Время обработки: {result.processing_time:.1f} сек\n"
                
                if result.words:
                    result_text += f"\nСлова: {len(result.words)}\n"
                    for i, word in enumerate(result.words[:10]):  # Показываем первые 10 слов
                        result_text += f"  {i+1}. {word.get('word', '')}\n"
                    if len(result.words) > 10:
                        result_text += f"  ... и еще {len(result.words) - 10} слов\n"
            else:
                result_text = "❌ РАСПОЗНАНИЕ НЕ УДАЛОСЬ\n\n"
                result_text += "Возможные причины:\n"
                result_text += "1. Аудиофайл поврежден\n"
                result_text += "2. В файле нет речи\n"
                result_text += "3. Модель не загружена корректно\n"
                result_text += "4. Неправильный формат файла\n"
            
            self.test_result_text.insert(tk.END, result_text)
            self.status_var.set("✅ Тест завершен")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка теста распознавания: {e}")
    
    def export_data(self):
        """Экспорт данных"""
        try:
            # Получаем список записей для экспорта
            recordings = self._get_recordings_for_export()
            
            if not recordings:
                messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
                return
            
            # Запрашиваем место сохранения
            format_type = self.export_format.get()
            
            if format_type == "csv":
                ext = ".csv"
                filetypes = [("CSV файлы", "*.csv")]
            elif format_type == "json":
                ext = ".json"
                filetypes = [("JSON файлы", "*.json")]
            elif format_type == "excel":
                ext = ".xlsx"
                filetypes = [("Excel файлы", "*.xlsx")]
            else:  # all
                ext = ".zip"
                filetypes = [("ZIP архив", "*.zip")]
            
            filename = filedialog.asksaveasfilename(
                defaultextension=ext,
                filetypes=filetypes,
                initialfile=f"sound_isolation_export{ext}"
            )
            
            if not filename:
                return
            
            # Экспортируем данные
            self._perform_export(recordings, filename, format_type)
            
            messagebox.showinfo("Успех", f"Данные экспортированы:\n{filename}")
            self.status_var.set("📁 Данные экспортированы")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка экспорта: {e}")
    
    def _get_recordings_for_export(self):
        """Получить записи для экспорта"""
        recordings = []
        selection_type = self.export_selection.get()
        
        try:
            # Сканируем папку recordings
            for file in os.listdir(self.recordings_folder):
                if file.endswith('_metadata.json'):
                    metadata_path = os.path.join(self.recordings_folder, file)
                    try:
                        with open(metadata_path, 'r', encoding='utf-8') as f:
                            metadata = json.load(f)
                            
                            # Проверяем фильтр по времени
                            if selection_type == "week":
                                timestamp = metadata.get('timestamp', '')
                                if timestamp:
                                    try:
                                        record_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                                        week_ago = datetime.now() - timedelta(days=7)
                                        if record_date < week_ago:
                                            continue
                                    except:
                                        pass
                            
                            recordings.append(metadata)
                            
                    except Exception as e:
                        print(f"Ошибка чтения {file}: {e}")
            
        except Exception as e:
            print(f"Ошибка получения записей: {e}")
        
        return recordings
    
    def _perform_export(self, recordings, filename, format_type):
        """Выполнить экспорт данных"""
        if format_type == "csv":
            self._export_to_csv(recordings, filename)
        elif format_type == "json":
            self._export_to_json(recordings, filename)
        elif format_type == "excel":
            self._export_to_excel(recordings, filename)
        elif format_type == "all":
            self._export_all_formats(recordings, filename)
    
    def _export_to_csv(self, recordings, filename):
        """Экспорт в CSV"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Заголовки
            writer.writerow([
                'test_name', 'timestamp', 'duration', 'sample_rate',
                'outside_samples', 'inside_samples', 'analysis_ready'
            ])
            
            # Данные
            for rec in recordings:
                writer.writerow([
                    rec.get('test_name', ''),
                    rec.get('timestamp', ''),
                    rec.get('duration', 0),
                    rec.get('sample_rate', 0),
                    rec.get('files', {}).get('outside', {}).get('samples', 0),
                    rec.get('files', {}).get('inside', {}).get('samples', 0),
                    'Да' if rec.get('analysis_ready', False) else 'Нет'
                ])
    
    def _export_to_json(self, recordings, filename):
        """Экспорт в JSON"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(recordings, f, ensure_ascii=False, indent=2)
    
    def _export_to_excel(self, recordings, filename):
        """Экспорт в Excel"""
        try:
            import pandas as pd
            
            # Преобразуем в DataFrame
            data = []
            for rec in recordings:
                data.append({
                    'test_name': rec.get('test_name', ''),
                    'timestamp': rec.get('timestamp', ''),
                    'duration': rec.get('duration', 0),
                    'sample_rate': rec.get('sample_rate', 0),
                    'outside_samples': rec.get('files', {}).get('outside', {}).get('samples', 0),
                    'inside_samples': rec.get('files', {}).get('inside', {}).get('samples', 0),
                    'analysis_ready': rec.get('analysis_ready', False)
                })
            
            df = pd.DataFrame(data)
            df.to_excel(filename, index=False)
            
        except ImportError:
            messagebox.showerror("Ошибка", 
                "Для экспорта в Excel установите pandas:\n"
                "pip install pandas openpyxl")
    
    def _export_all_formats(self, recordings, filename):
        """Экспорт во всех форматах"""
        import zipfile
        import tempfile
        import os
        
        # Создаем временную папку
        with tempfile.TemporaryDirectory() as temp_dir:
            # Экспортируем во все форматы
            csv_file = os.path.join(temp_dir, "data.csv")
            json_file = os.path.join(temp_dir, "data.json")
            excel_file = os.path.join(temp_dir, "data.xlsx")
            readme_file = os.path.join(temp_dir, "README.txt")
            
            self._export_to_csv(recordings, csv_file)
            self._export_to_json(recordings, json_file)
            
            try:
                self._export_to_excel(recordings, excel_file)
            except:
                # Если Excel не доступен, создаем заглушку
                with open(excel_file, 'w') as f:
                    f.write("Excel export requires pandas and openpyxl\n")
            
            # Создаем README
            with open(readme_file, 'w', encoding='utf-8') as f:
                f.write("ЭКСПОРТ ДАННЫХ ТЕСТЕРА ЗВУКОИЗОЛЯЦИИ\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Количество записей: {len(recordings)}\n\n")
                f.write("ФАЙЛЫ:\n")
                f.write("1. data.csv - данные в формате CSV\n")
                f.write("2. data.json - данные в формате JSON\n")
                f.write("3. data.xlsx - данные в формате Excel (если доступно)\n\n")
                f.write("ДЛЯ ИМПОРТА В ДИПЛОМНУЮ РАБОТУ:\n")
                f.write("• Используйте Excel для графиков\n")
                f.write("• Используйте CSV для статистического анализа\n")
                f.write("• Используйте JSON для программирования\n")
            
            # Создаем ZIP архив
            with zipfile.ZipFile(filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file in [csv_file, json_file, excel_file, readme_file]:
                    if os.path.exists(file):
                        zipf.write(file, os.path.basename(file))
    
    def load_config(self):
        """Загрузить конфигурацию"""
        try:
            config_file = "config.json"
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # Восстанавливаем настройки
                if 'last_engine' in config:
                    # Пытаемся установить последний движок
                    pass
                
                print("✅ Конфигурация загружена")
                
        except Exception as e:
            print(f"⚠️ Ошибка загрузки конфигурации: {e}")
    
    def save_config(self):
        """Сохранить конфигурацию"""
        try:
            config = {
                'last_engine': self.current_engine.value if self.current_engine else None,
                'save_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            with open("config.json", 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            print("✅ Конфигурация сохранена")
            
        except Exception as e:
            print(f"⚠️ Ошибка сохранения конфигурации: {e}")
    
    def on_closing(self):
        """Обработчик закрытия окна"""
        try:
            # Останавливаем мониторинг
            self.monitoring_active = False
            
            # Сохраняем конфигурацию
            self.save_config()
            
            # Очищаем ресурсы
            if hasattr(self, 'audio_core'):
                self.audio_core.cleanup()
            
            # Закрываем приложение
            self.root.destroy()
            
        except Exception as e:
            print(f"Ошибка при закрытии: {e}")
            self.root.destroy()

def main():
    """Главная функция"""
    try:
        root = tk.Tk()
        app = AdvancedSoundTester(root)
        
        # Обработчик закрытия окна
        root.protocol("WM_DELETE_WINDOW", app.on_closing)
        
        root.mainloop()
        
    except Exception as e:
        import traceback
        error_msg = f"Критическая ошибка запуска:\n\n{str(e)}\n\n"
        error_msg += "Трассировка:\n"
        error_msg += traceback.format_exc()
        
        print(error_msg)
        
        # Пытаемся показать сообщение об ошибке
        try:
            tk.Tk().withdraw()  # Скрываем основное окно
            messagebox.showerror("Критическая ошибка", 
                f"Ошибка запуска:\n\n{str(e)[:200]}...\n\n"
                f"Проверьте:\n"
                f"1. Все файлы в одной папке\n"
                f"2. Установлены зависимости: pip install -r requirements.txt\n"
                f"3. Проверьте консоль для подробной информации")
        except:
            pass

if __name__ == "__main__":
    main()